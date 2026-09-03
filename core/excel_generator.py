import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Protection

class ExcelGenerator:

    def __init__(self, language):
        self.language = language

    def generate(
        self,
        students,
        student_groups,
        assignments,
        practice_questions,
        questions_per_student,
        theory_max,
        defense_max,
        defense_percentage,
        theory_percentage,
        total_score,
        course,
        module,
        term,
        course_language
    ):
        
        """
        Generate the Excel assignment workbook.
        """

        # Create workbook and worksheets
        workbook = Workbook()

        workbook.properties.creator = "Ing. Lucio Buitrón"
        workbook.properties.lastModifiedBy = "Ing. Lucio Buitrón"
        workbook.properties.title = "Question Assignment Tool"
        workbook.properties.subject = "Question Assignment"
        workbook.properties.description = (
            "Assignment file generated with Question Assignment Tool v3.0"
        )
        workbook.properties.keywords = (
            "Question Assignment, Education, Assessment, Excel"
        )
        workbook.properties.category = "Education"

        worksheet = workbook.active
        worksheet.title = "Theoretical Questions"

        practice_worksheet = None

        if practice_questions:
            practice_worksheet = workbook.create_sheet("Practice Questions")

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78"
        )

        light_blue_fill = PatternFill(
            fill_type="solid",
            start_color="EAF4FF"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        medium_bottom_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="medium")
        )

        center_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        left_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        # =====================================================
        # Grade References - Top Section
        # =====================================================

        defense_grade_reference = [
            ("EXCELLENT", defense_max),
            ("GOOD", defense_max * 0.75),
            ("FAIR", defense_max * 0.50),
            ("POOR", defense_max * 0.25)
        ]

        theory_grade_reference = [
            ("EXCELLENT", theory_max),
            ("GOOD", theory_max * 0.75),
            ("FAIR", theory_max * 0.50),
            ("POOR", theory_max * 0.25)
        ]

        # -----------------------------------------------------
        # Defense Grade Reference - A1:B6
        # -----------------------------------------------------

        worksheet.cell(
            row=1,
            column=1
        ).value = f"Defense Grade Reference ({defense_percentage:.0f}%)"

        worksheet.cell(
            row=1,
            column=1
        ).font = Font(
            bold=True,
            size=10
        )

        worksheet.cell(row=2, column=1).value = "Grade"
        worksheet.cell(row=2, column=2).value = "Points"

        for cell in worksheet[2]:
            if cell.column <= 2:
                cell.fill = header_fill
                cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                    size=8
                )
                cell.alignment = header_alignment

        for index, (grade, points) in enumerate(
            defense_grade_reference,
            start=3
        ):
            worksheet.cell(row=index, column=1).value = grade
            worksheet.cell(row=index, column=2).value = points

            worksheet.cell(
                row=index,
                column=1
            ).alignment = center_alignment

            worksheet.cell(
                row=index,
                column=2
            ).alignment = left_alignment

            worksheet.cell(
                row=index,
                column=1
            ).border = thin_border

            worksheet.cell(
                row=index,
                column=2
            ).border = thin_border

            worksheet.cell(
                row=index,
                column=1
            ).font = Font(size=8)

            worksheet.cell(
                row=index,
                column=2
            ).font = Font(size=8)

        # -----------------------------------------------------
        # Theoretical Question Grade Reference - D1:E6
        # -----------------------------------------------------

        worksheet.cell(
            row=1,
            column=4
        ).value = (
            f"Theoretical Question Grade Reference "
            f"({theory_percentage:.0f}%)"
        )

        worksheet.cell(
            row=1,
            column=4
        ).font = Font(
            bold=True,
            size=10
        )

        worksheet.cell(row=2, column=4).value = "Grade"
        worksheet.cell(row=2, column=5).value = "Points"

        for cell in worksheet[2]:
            if cell.column in (4, 5):
                cell.fill = header_fill
                cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                    size=8
                )
                cell.alignment = header_alignment

        for index, (grade, points) in enumerate(
            theory_grade_reference,
            start=3
        ):
            worksheet.cell(row=index, column=4).value = grade
            worksheet.cell(row=index, column=5).value = points

            worksheet.cell(
                row=index,
                column=4
            ).alignment = center_alignment

            worksheet.cell(
                row=index,
                column=5
            ).alignment = left_alignment

            worksheet.cell(
                row=index,
                column=4
            ).border = thin_border

            worksheet.cell(
                row=index,
                column=5
            ).border = thin_border

            worksheet.cell(
                row=index,
                column=4
            ).font = Font(size=8)

            worksheet.cell(
                row=index,
                column=5
            ).font = Font(size=8)

            worksheet.cell(
                row=8,
                column=1
            ).value = "TOTAL GRADE"

            worksheet.cell(
                row=8,
                column=2
            ).value = total_score

            worksheet.cell(
                row=8,
                column=1
            ).font = Font(
                bold=True,
                color="FFFFFF",
                size=8
            )

            worksheet.cell(
                row=8,
                column=2
            ).font = Font(
                bold=True,
                color="FFFFFF",
                size=12
            )

            worksheet.cell(row=8, column=1).fill = header_fill
            worksheet.cell(row=8, column=2).fill = header_fill

        # Assignments table starts after Grade Reference
        worksheet.cell(row=10, column=1).value = "Group"
        worksheet.cell(row=10, column=2).value = "#"
        worksheet.cell(row=10, column=3).value = "Student"
        worksheet.cell(row=10, column=4).value = "Defense Grade"
        worksheet.cell(row=10, column=5).value = "Defense Grade Value"
        worksheet.cell(row=10, column=6).value = "Theoretical Question"
        worksheet.cell(row=10, column=7).value = "Theoretical Question Grade"
        worksheet.cell(row=10, column=8).value = "Theoretical Question Grade Value"
        worksheet.cell(row=10, column=9).value = "Practice Question Grade"
        worksheet.cell(row=10, column=10).value = "FINAL TOTAL GRADE"
        worksheet.cell(row=10, column=11).value = "Comments"

        # =====================================================
        # Practice Questions Worksheet
        # =====================================================

        if practice_questions:

            # -------------------------------------------------
            # Practice Grade Reference
            # -------------------------------------------------

            practice_max = total_score * 0.10
            practice_partial = total_score * 0.05

            practice_grade_reference = [
                ("RESOLVED", practice_max),
                ("PARTIAL", practice_partial),
                ("NOT RESOLVED", 0)
            ]

            practice_worksheet.cell(
                row=1,
                column=1
            ).value = "Practice Question Grade Reference (EXTRA)"

            practice_worksheet.cell(
                row=1,
                column=1
            ).font = Font(
                bold=True,
                size=10
            )

            practice_worksheet.cell(
                row=2,
                column=1
            ).value = "Grade"

            practice_worksheet.cell(
                row=2,
                column=2
            ).value = "Points"

            practice_worksheet.cell(
                row=2,
                column=3
            ).value = "Explanation"

            for cell in practice_worksheet[2]:
                if cell.column <= 3:
                    cell.fill = header_fill
                    cell.font = Font(
                        bold=True,
                        color="FFFFFF",
                        size=8
                    )
                    cell.alignment = header_alignment

            for index, (grade, points) in enumerate(
                practice_grade_reference,
                start=3
            ):
                practice_worksheet.cell(
                    row=index,
                    column=1
                ).value = grade

                practice_worksheet.cell(
                    row=index,
                    column=2
                ).value = points

                practice_worksheet.cell(
                    row=index,
                    column=3
                ).value = (
                    f"{10 if grade == 'RESOLVED' else 5 if grade == 'PARTIAL' else 0}% "
                    f"of Total Grade: {total_score:g} points"
                )

                practice_worksheet.cell(
                    row=index,
                    column=1
                ).alignment = center_alignment

                practice_worksheet.cell(
                    row=index,
                    column=2
                ).alignment = left_alignment

                practice_worksheet.cell(
                    row=index,
                    column=1
                ).border = thin_border

                practice_worksheet.cell(
                    row=index,
                    column=2
                ).border = thin_border

                practice_worksheet.cell(
                    row=index,
                    column=1
                ).font = Font(size=8)

                practice_worksheet.cell(
                    row=index,
                    column=2
                ).font = Font(size=8)

                practice_worksheet.cell(
                    row=index,
                    column=3
                ).alignment = left_alignment

                practice_worksheet.cell(
                    row=index,
                    column=3
                ).border = thin_border

                practice_worksheet.cell(
                    row=index,
                    column=3
                ).font = Font(size=8)

            # -------------------------------------------------
            # Practice Questions Table
            # -------------------------------------------------

            practice_worksheet.cell(
                row=10,
                column=1
            ).value = "Group"

            practice_worksheet.cell(
                row=10,
                column=2
            ).value = "#"

            practice_worksheet.cell(
                row=10,
                column=3
            ).value = "Student"

            practice_worksheet.cell(
                row=10,
                column=4
            ).value = "Practice Question"

            practice_worksheet.cell(
                row=10,
                column=5
            ).value = "Practice Question Grade"

            practice_worksheet.cell(
                row=10,
                column=6
            ).value = "Practice Question Grade Value"

            practice_worksheet.cell(
                row=10,
                column=7
            ).value = "Comments"

            # -------------------------------------------------
            # Header Merge
            # -------------------------------------------------

            practice_worksheet.merge_cells(
                start_row=10,
                start_column=5,
                end_row=10,
                end_column=6
            )

            # -------------------------------------------------
            # Practice Question Validation
            # -------------------------------------------------

            practice_question_list_sheet = workbook.create_sheet(
                "_PracticeLists"
            )

            for index, practice_question in enumerate(
                practice_questions,
                start=1
            ):
                practice_question_list_sheet.cell(
                    row=index,
                    column=1
                ).value = practice_question

            practice_question_list_sheet.sheet_state = "hidden"

            practice_question_validation = DataValidation(
                type="list",
                formula1=(
                    f"'_PracticeLists'!$A$1:$A$"
                    f"{len(practice_questions)}"
                ),
                allow_blank=True
            )

            practice_grade_validation = DataValidation(
                type="list",
                formula1='"RESOLVED,PARTIAL,NOT RESOLVED"',
                allow_blank=True
            )
            practice_grade_validation.errorStyle = "stop"
            practice_grade_validation.showErrorMessage = True
            practice_grade_validation.errorTitle = "Invalid Value"
            practice_grade_validation.error = "Please select a value from the dropdown list."

            practice_worksheet.add_data_validation(
                practice_question_validation
            )

            practice_worksheet.add_data_validation(
                practice_grade_validation
            )

            # -------------------------------------------------
            # Populate Practice Questions
            # -------------------------------------------------

            row = 11
            student_number = 1
            current_group = None
            group_start_row = None
            group_number = 0
            group_end_rows = []

            for student in students:

                start_row = row
                group = student_groups[student]

                if group != current_group:

                    if (
                        current_group is not None
                        and group_start_row < start_row
                    ):
                        # Group merge
                        practice_worksheet.merge_cells(
                            start_row=group_start_row,
                            start_column=1,
                            end_row=start_row - 1,
                            end_column=1
                        )

                        # Practice Question merge
                        practice_worksheet.merge_cells(
                            start_row=group_start_row,
                            start_column=4,
                            end_row=start_row - 1,
                            end_column=4
                        )

                        # Group colors
                        if group_number % 2 == 1:
                            for r in range(
                                group_start_row,
                                start_row
                            ):
                                practice_worksheet.cell(
                                    row=r,
                                    column=1
                                ).fill = light_blue_fill

                            practice_worksheet.cell(
                                row=group_start_row,
                                column=4
                            ).fill = light_blue_fill

                    group_end_rows.append(start_row - 1)
                    current_group = group
                    group_start_row = start_row
                    group_number += 1

                # Group
                practice_worksheet.cell(
                    row=start_row,
                    column=1
                ).value = group

                practice_worksheet.cell(
                    row=start_row,
                    column=1
                ).alignment = center_alignment

                # Student #
                practice_worksheet.cell(
                    row=start_row,
                    column=2
                ).value = student_number

                practice_worksheet.cell(
                    row=start_row,
                    column=2
                ).alignment = center_alignment

                # Student
                practice_worksheet.cell(
                    row=start_row,
                    column=3
                ).value = student

                practice_worksheet.cell(
                    row=start_row,
                    column=3
                ).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True
                )

                # Practice Question
                practice_worksheet.cell(
                    row=start_row,
                    column=4
                ).value = ""

                practice_worksheet.cell(
                    row=start_row,
                    column=4
                ).alignment = left_alignment

                practice_question_validation.add(
                    practice_worksheet.cell(
                        row=start_row,
                        column=4
                    )
                )
                practice_question_validation.errorStyle = "stop"
                practice_question_validation.showErrorMessage = True
                practice_question_validation.errorTitle = "Invalid Value"
                practice_question_validation.error = "Please select a value from the dropdown list."

                # Practice Question Grade
                practice_worksheet.cell(
                    row=start_row,
                    column=5
                ).value = ""

                practice_worksheet.cell(
                    row=start_row,
                    column=5
                ).alignment = center_alignment

                practice_grade_validation.add(
                    practice_worksheet.cell(
                        row=start_row,
                        column=5
                    )
                )

                # Practice Question Grade Value
                practice_worksheet.cell(
                    row=start_row,
                    column=6
                ).value = (
                    f'=IF(E{start_row}="RESOLVED",'
                    f'{practice_max},'
                    f'IF(E{start_row}="PARTIAL",'
                    f'{practice_partial},'
                    f'IF(E{start_row}="NOT RESOLVED",'
                    f'0,"")))'
                )

                practice_worksheet.cell(
                    row=start_row,
                    column=6
                ).alignment = center_alignment

                # Comments
                practice_worksheet.cell(
                    row=start_row,
                    column=7
                ).value = ""

                practice_worksheet.cell(
                    row=start_row,
                    column=7
                ).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True
                )

                # Student colors
                if student_number % 2 == 1:
                    for column in (2, 3, 5, 6, 7):
                        practice_worksheet.cell(
                            row=start_row,
                            column=column
                        ).fill = light_blue_fill

                student_number += 1
                row += 1

        if practice_questions:
            # -------------------------------------------------
            # Final Group Merge
            # -------------------------------------------------

            if current_group is not None and group_start_row < row:

                # Group merge
                practice_worksheet.merge_cells(
                    start_row=group_start_row,
                    start_column=1,
                    end_row=row - 1,
                    end_column=1
                )

                # Practice Question merge
                practice_worksheet.merge_cells(
                    start_row=group_start_row,
                    start_column=4,
                    end_row=row - 1,
                    end_column=4
                )

                # Group colors
                if group_number % 2 == 1:
                    for r in range(
                        group_start_row,
                        row
                    ):
                        practice_worksheet.cell(
                            row=r,
                            column=1
                        ).fill = light_blue_fill

                    practice_worksheet.cell(
                        row=group_start_row,
                        column=4
                    ).fill = light_blue_fill

                # -------------------------------------------------
                # Headers
                # -------------------------------------------------

                for cell in practice_worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                if current_group is not None:
                    group_end_rows.append(row - 1)

                # -------------------------------------------------
                # Borders
                # -------------------------------------------------

                for row_cells in practice_worksheet.iter_rows(
                    min_row=10
                ):
                    for cell in row_cells:
                        cell.border = thin_border

                for group_end_row in group_end_rows:
                    for column in range(1, 8):
                        practice_worksheet.cell(
                            row=group_end_row,
                            column=column
                        ).border = medium_bottom_border

                # -------------------------------------------------
                # Column Widths
                # -------------------------------------------------

                practice_worksheet.column_dimensions["A"].width = 10
                practice_worksheet.column_dimensions["B"].width = 5
                practice_worksheet.column_dimensions["C"].width = 40
                practice_worksheet.column_dimensions["D"].width = 90
                practice_worksheet.column_dimensions["E"].width = 15
                practice_worksheet.column_dimensions["F"].width = 10
                practice_worksheet.column_dimensions["G"].width = 60

                # -------------------------------------------------
                # Headers
                # -------------------------------------------------

                for cell in practice_worksheet[10]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

        for cell in worksheet[10]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        worksheet.cell(row=10, column=1).alignment = header_alignment
        worksheet.cell(row=10, column=2).alignment = header_alignment
        worksheet.cell(row=10, column=3).alignment = header_alignment
        worksheet.cell(row=10, column=4).alignment = header_alignment

        for column in range(5, 9):
            worksheet.cell(
                row=10,
                column=column
            ).fill = header_fill

            worksheet.cell(
                row=10,
                column=column
            ).font = header_font

        worksheet.cell(
            row=10,
            column=7
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
             
        grade_validation = DataValidation(
            type="list",
            formula1='"Not Graded,Excellent,Good,Fair,Poor"',
            allow_blank=True
        )
        grade_validation.errorStyle = "stop"
        grade_validation.showErrorMessage = True
        grade_validation.errorTitle = "Invalid Value"
        grade_validation.error = "Please select a value from the dropdown list."

        defense_grade_validation = DataValidation(
            type="list",
            formula1='"Not Graded,Excellent,Good,Fair,Poor"',
            allow_blank=True
        )
        defense_grade_validation.errorStyle = "stop"
        defense_grade_validation.showErrorMessage = True
        defense_grade_validation.errorTitle = "Invalid Value"
        defense_grade_validation.error = "Please select a value from the dropdown list."

        worksheet.add_data_validation(grade_validation)
        worksheet.add_data_validation(defense_grade_validation)
        
        worksheet.merge_cells(
            start_row=10,
            start_column=4,
            end_row=10,
            end_column=5
        )

        worksheet.merge_cells(
            start_row=10,
            start_column=7,
            end_row=10,
            end_column=8
        )

        # Populate Theoretical Questions
        row = 11
        student_number = 1
        current_group = None
        group_start_row = None
        group_number = 0
        group_end_rows = []

        for student in students:
            assigned_questions = assignments[student]
            start_row = row
            group = student_groups[student]

            if group != current_group:
                if current_group is not None and group_start_row < start_row:
                    worksheet.merge_cells(
                        start_row=group_start_row,
                        start_column=1,
                        end_row=start_row - 1,
                        end_column=1
                    )

                    if group_number % 2 == 1:
                        for r in range(group_start_row, start_row):
                            worksheet.cell(
                                row=r,
                                column=1
                            ).fill = light_blue_fill

                group_end_rows.append(start_row - 1)
                current_group = group
                group_start_row = start_row
                group_number += 1

            # Theoretical Questions
            for question in assigned_questions:
                worksheet.cell(row=row, column=6).value = question
                worksheet.cell(
                    row=row,
                    column=6
                ).alignment = left_alignment

                # Theoretical Question Grade
                worksheet.cell(row=row, column=7).value = ""
                worksheet.cell(
                    row=row,
                    column=7
                ).alignment = center_alignment

                grade_validation.add(
                    worksheet.cell(row=row, column=7)
                )

                row += 1

            end_row = row - 1

            # Group
            worksheet.cell(row=start_row, column=1).value = group
            worksheet.cell(
                row=start_row,
                column=1
            ).alignment = center_alignment

            # Student #
            worksheet.cell(row=start_row, column=2).value = student_number
            worksheet.cell(
                row=start_row,
                column=2
            ).alignment = center_alignment

            # Student
            worksheet.cell(row=start_row, column=3).value = student
            worksheet.cell(
                row=start_row,
                column=3
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

            # Defense Grade
            worksheet.cell(row=start_row, column=4).value = ""
            worksheet.cell(
                row=start_row,
                column=4
            ).alignment = center_alignment

            defense_grade_validation.add(
                worksheet.cell(row=start_row, column=4)
            )

            # Defense Grade Value
            worksheet.cell(
                row=start_row,
                column=5
            ).value = (
                f'=IF(D{start_row}="EXCELLENT",{defense_max},'
                f'IF(D{start_row}="GOOD",{defense_max * 0.75},'
                f'IF(D{start_row}="FAIR",{defense_max * 0.50},'
                f'IF(D{start_row}="POOR",{defense_max * 0.25},""))))'
            )

            worksheet.cell(
                row=start_row,
                column=5
            ).alignment = center_alignment

            # Theoretical Question Grade Value
            worksheet.cell(
                row=start_row,
                column=8
            ).value = (
                f'=IFERROR(('
                f'COUNTIF(G{start_row}:G{end_row},"EXCELLENT")*{theory_max}+'
                f'COUNTIF(G{start_row}:G{end_row},"GOOD")*{theory_max * 0.75}+'
                f'COUNTIF(G{start_row}:G{end_row},"FAIR")*{theory_max * 0.50}+'
                f'COUNTIF(G{start_row}:G{end_row},"POOR")*{theory_max * 0.25}'
                f')/'
                f'('
                f'COUNTIF(G{start_row}:G{end_row},"EXCELLENT")+'
                f'COUNTIF(G{start_row}:G{end_row},"GOOD")+'
                f'COUNTIF(G{start_row}:G{end_row},"FAIR")+'
                f'COUNTIF(G{start_row}:G{end_row},"POOR")'
                f'),"")'
            )

            worksheet.cell(
                row=start_row,
                column=8
            ).alignment = center_alignment

            # Practice Question Grade
            if practice_questions:
                worksheet.cell(
                    row=start_row,
                    column=9
                ).value = (
                    f"='Practice Questions'!F{student_number + 10}"
                )
            else:
                worksheet.cell(
                    row=start_row,
                    column=9
                ).value = ""

            worksheet.cell(
                row=start_row,
                column=9
            ).alignment = center_alignment

            # FINAL TOTAL GRADE
            worksheet.cell(
                row=start_row,
                column=10
            ).value = (
                f'=IFERROR(E{start_row}+H{start_row}+'
                f'IF(I{start_row}="",0,I{start_row}),"")'
            )

            worksheet.cell(
                row=start_row,
                column=10
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            worksheet.cell(
                row=start_row,
                column=10
            ).font = Font(
                bold=True,
                size=13
            )

            # Comments
            worksheet.cell(row=start_row, column=11).value = ""
            worksheet.cell(
                row=start_row,
                column=11
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )

            # Merge student-level columns
            if end_row > start_row:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=2,
                    end_row=end_row,
                    end_column=2
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=3,
                    end_row=end_row,
                    end_column=3
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=4,
                    end_row=end_row,
                    end_column=4
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=5,
                    end_row=end_row,
                    end_column=5
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=8,
                    end_row=end_row,
                    end_column=8
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=9,
                    end_row=end_row,
                    end_column=9
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=10,
                    end_row=end_row,
                    end_column=10
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=11,
                    end_row=end_row,
                    end_column=11
                )

            # Student colors
            if student_number % 2 == 1:
                for r in range(start_row, end_row + 1):
                    for c in range(2, 12):
                        worksheet.cell(
                            row=r,
                            column=c
                        ).fill = light_blue_fill

            student_number += 1

        # Final Group merge
        if current_group is not None and group_start_row < row:
            worksheet.merge_cells(
                start_row=group_start_row,
                start_column=1,
                end_row=row - 1,
                end_column=1
            )

            if group_number % 2 == 1:
                for r in range(group_start_row, row):
                    worksheet.cell(
                        row=r,
                        column=1
                    ).fill = light_blue_fill

        if current_group is not None:
            group_end_rows.append(row - 1)
        
        # Column widths
        worksheet.column_dimensions["A"].width = 10
        worksheet.column_dimensions["B"].width = 7
        worksheet.column_dimensions["C"].width = 40
        worksheet.column_dimensions["D"].width = 12
        worksheet.column_dimensions["E"].width = 10
        worksheet.column_dimensions["F"].width = 90
        worksheet.column_dimensions["G"].width = 12
        worksheet.column_dimensions["H"].width = 10
        worksheet.column_dimensions["I"].width = 10
        worksheet.column_dimensions["J"].width = 10
        worksheet.column_dimensions["K"].width = 50

        # Apply final worksheet formatting
        for row in worksheet.iter_rows(min_row=10):
            for cell in row:
                cell.border = thin_border

            for group_end_row in group_end_rows:
                for column in range(1, 12):
                    worksheet.cell(
                        row=group_end_row,
                        column=column
                    ).border = medium_bottom_border

            worksheet.cell(
                row=10,
                column=2
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            worksheet.cell(
                row=10,
                column=3
            ).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # =====================================================
        # Save Excel
        # =====================================================

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        course = course.replace(" ", "")
        module = module.replace(" ", "")
        filename = (
            f"{course}_"
            f"{module}_"
            f"Term{term}_"
            f"{course_language}_"
            f"{timestamp}.xlsx"
        )

        output_folder = "output"
        os.makedirs(output_folder, exist_ok=True)
        filepath = os.path.join(output_folder, filename)
        filepath = os.path.abspath(filepath)

        # Protect Theoretical Questions worksheet
        worksheet.protection.sheet = True

        if practice_questions:
            for r in range(11, practice_worksheet.max_row + 1):
                for c in range(1, 8):
                    practice_worksheet.cell(
                        row=r,
                        column=c
                    ).protection = Protection(locked=True)

                for c in [1, 4, 5, 7]:
                    practice_worksheet.cell(
                        row=r,
                        column=c
                    ).protection = Protection(locked=False)

            practice_worksheet.protection.sheet = True
        
        # Cell protection
        for r in range(11, worksheet.max_row + 1):
            for c in range(1, 12):
                worksheet.cell(
                    row=r,
                    column=c
                ).protection = Protection(locked=True)

            editable_columns = [1, 4, 7, 11]

            if not practice_questions:
                editable_columns.append(9)

            for c in editable_columns:
                worksheet.cell(
                    row=r,
                    column=c
                ).protection = Protection(locked=False)

        workbook.save(filepath)
        return filepath

# === QAT V3 STABLE CHECKPOINT ===
# Protection + Data Validation in all cells and tabs tested and working.
# Do not modify this configuration without a backup/commit.