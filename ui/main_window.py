import random
import json
import os
import customtkinter as ctk
import tkinter as tk

from core.language_manager import LanguageManager
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from datetime import datetime
from tkinter import messagebox


class MainWindow(ctk.CTk):

    def __init__(self):
        super().__init__()

        self.load_configuration()
        self.language = LanguageManager(
            self.config_data["general"]["application_language"]
        )
        self.create_menu()
        self.configure_window()
        self.create_widgets()

    def create_menu(self):

        self.menu_bar = tk.Menu(self)

        self.help_menu = tk.Menu(
            self.menu_bar,
            tearoff=0
        )

        self.menu_bar.add_cascade(
            label=self.language.get("help"),
            menu=self.help_menu
        )

        self.help_menu.add_command(
            label=self.language.get("instructions"),
            command=self.show_instructions
        )

        self.help_menu.add_separator()

        self.help_menu.add_command(
            label=self.language.get("about"),
            command=self.show_about
        )

        self.config(menu=self.menu_bar)

    def load_configuration(self):

        with open("config.json", "r", encoding="utf-8") as file:
            self.config_data = json.load(file)

    def configure_window(self):

        app = self.config_data["application"]
        window = self.config_data["window"]

        self.title(f'{self.language.get("title")} v{app["version"]}')

        self.geometry(f'{window["width"]}x{window["height"]}')

        self.minsize(
            window["min_width"],
            window["min_height"]
        )

        self.resizable(
            window["resizable"],
            window["resizable"]
        )

    def create_widgets(self):

        lists = self.config_data["lists"]

        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        for column in range(5):
            self.main_frame.grid_columnconfigure(column, weight=1)

        # =====================================================
        # Title
        # =====================================================

        self.lbl_title = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("title"),
            font=("Segoe UI", 24, "bold")
        )

        self.lbl_title.grid(
            row=0,
            column=0,
            columnspan=5,
            pady=(10, 20)
        )

        # =====================================================
        # Application Language
        # =====================================================

        self.lbl_application_language = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("application_language")
        )

        self.lbl_application_language.grid(
            row=1,
            column=0,
            columnspan=5,
            pady=(0, 5)
        )

        self.cmb_application_language = ctk.CTkComboBox(
            self.main_frame,
            values=list(lists["application_languages"].keys()),
            command=self.change_language
        )

        language_code = self.config_data["general"]["application_language"]

        for name, code in lists["application_languages"].items():
            if code == language_code:
                self.cmb_application_language.set(name)
                break

        self.cmb_application_language.grid(
            row=2,
            column=2,
            padx=5,
            pady=(0, 20),
            sticky="ew"
        )

        # =====================================================
        # Course
        # =====================================================

        self.lbl_course = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("course")
        )

        self.lbl_course.grid(row=3, column=0, padx=5, sticky="w")

        self.cmb_course = ctk.CTkComboBox(
            self.main_frame,
            values=lists["courses"]
        )

        self.cmb_course.grid(row=4, column=0, padx=5, sticky="ew")

        # =====================================================
        # Module
        # =====================================================

        self.lbl_module = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("module")
        )

        self.lbl_module.grid(row=3, column=1, padx=5, sticky="w")

        self.cmb_module = ctk.CTkComboBox(
            self.main_frame,
            values=lists["modules"]
        )

        self.cmb_module.grid(row=4, column=1, padx=5, sticky="ew")

        # =====================================================
        # Term
        # =====================================================

        self.lbl_term = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("term")
        )

        self.lbl_term.grid(row=3, column=2, padx=5, sticky="w")

        self.cmb_term = ctk.CTkComboBox(
            self.main_frame,
            values=lists["terms"]
        )

        self.cmb_term.grid(row=4, column=2, padx=5, sticky="ew")

        # =====================================================
        # Year
        # =====================================================

        self.lbl_year = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("year")
        )

        self.lbl_year.grid(row=3, column=3, padx=5, sticky="w")

        self.cmb_year = ctk.CTkComboBox(
            self.main_frame,
            values=[str(year) for year in range(2026, 2036)]
        )

        self.cmb_year.set("2026")

        self.cmb_year.grid(row=4, column=3, padx=5, sticky="ew")

        # =====================================================
        # Course Language
        # =====================================================

        self.lbl_course_language = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("course_language")
        )

        self.lbl_course_language.grid(row=3, column=4, padx=5, sticky="w")

        self.cmb_course_language = ctk.CTkComboBox(
            self.main_frame,
            values=lists["course_languages"]
        )

        self.cmb_course_language.grid(row=4, column=4, padx=5, sticky="ew")

        # =====================================================
        # Student List
        # =====================================================

        self.lbl_students = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("student_list")
        )

        self.lbl_students.grid(
            row=5,
            column=0,
            columnspan=5,
            sticky="w",
            pady=(25, 5)
        )

        self.txt_students = ctk.CTkTextbox(
            self.main_frame,
            height=180
        )

        self.txt_students.grid(
            row=6,
            column=0,
            columnspan=5,
            sticky="nsew"
        )

        # =====================================================
        # Question Bank
        # =====================================================

        self.lbl_questions = ctk.CTkLabel(
            self.main_frame,
            text=self.language.get("question_bank")
        )

        self.lbl_questions.grid(
            row=7,
            column=0,
            columnspan=5,
            sticky="w",
            pady=(25, 5)
        )

        self.txt_questions = ctk.CTkTextbox(
            self.main_frame,
            height=180
        )

        self.txt_questions.grid(
            row=8,
            column=0,
            columnspan=5,
            sticky="nsew"
        )

        # =====================================================
        # Generate
        # =====================================================

        self.btn_generate = ctk.CTkButton(
            self.main_frame,
            text=self.language.get("generate_excel"),
            command=self.generate_excel
        )

        self.btn_generate.grid(
            row=9,
            column=0,
            columnspan=5,
            pady=25
        )

    
        # =====================================================
        # Exit
        # =====================================================

        self.btn_exit = ctk.CTkButton(
            self.main_frame,
            text=self.language.get("exit"),
            command=self.destroy
        )

        self.btn_exit.grid(
            row=12,
            column=0,
            columnspan=5,
            pady=(0, 10)
        )
    
    # =====================================================
    # Generate Excel
    # =====================================================
    def generate_excel(self):

        students = [
            student.strip()
            for student in self.txt_students.get("1.0", "end").splitlines()
            if student.strip()
        ]

        questions = [
            question.strip()
            for question in self.txt_questions.get("1.0", "end").splitlines()
            if question.strip()
        ]

        if not students:
            messagebox.showwarning(
                self.language.get("warning_title"),
                self.language.get("no_students_found")
            )
            return

        if not questions:
            messagebox.showwarning(
                self.language.get("warning_title"),
                self.language.get("no_questions_found")
            )
            return

        questions_per_student = self.config_data["general"]["questions_per_student"]

        required_questions = len(students) * questions_per_student

        if len(questions) < required_questions:
            messagebox.showwarning(
                self.language.get("warning_title"),
                self.language.get("not_enough_questions")
            )
            return

        available_questions = questions.copy()
        random.shuffle(available_questions)

        assignments = {}

        for student in students:

            assigned_questions = []

            for _ in range(questions_per_student):
                assigned_questions.append(available_questions.pop())

            assignments[student] = assigned_questions

        for student, assigned_questions in assignments.items():

            print(student)

            for question in assigned_questions:
                print(f"  - {question}")

        workbook = Workbook()

        workbook.properties.creator = "Ing. Lucio Buitrón"
        workbook.properties.lastModifiedBy = "Ing. Lucio Buitrón"
        workbook.properties.title = "Question Assignment Tool"
        workbook.properties.subject = "Question Assignment"
        workbook.properties.description = (
            "Assignment file generated with Question Assignment Tool v1.0"
        )
        workbook.properties.keywords = (
            "Question Assignment, Education, Assessment, Excel"
        )
        workbook.properties.category = "Education"

        worksheet = workbook.active
        worksheet.title = "Assignments"

        worksheet.append([
            "#",
            "Student",
            "Question",
            "Grade",
            "Comment"
        ])

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78"
        )

        light_blue_fill = PatternFill(
            fill_type="solid",
            start_color="EAF4FF"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        center_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        left_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        for row in range(2, worksheet.max_row + 1):

            worksheet.cell(row=row, column=1).alignment = center_alignment
            worksheet.cell(row=row, column=2).alignment = left_alignment
            worksheet.cell(row=row, column=3).alignment = left_alignment
            worksheet.cell(row=row, column=4).alignment = center_alignment
            worksheet.cell(row=row, column=5).alignment = left_alignment

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        grade_validation = DataValidation(
            type="list",
            formula1='"Not Graded,Excellent,Good,Fair,Poor"',
            allow_blank=True
        )

        worksheet.add_data_validation(grade_validation)

        row = 2
        student_number = 1

        for student, assigned_questions in assignments.items():

            start_row = row

            for question in assigned_questions:

                worksheet.cell(row=row, column=3).value = question
                worksheet.cell(row=row, column=4).value = ""
                grade_validation.add(
                    worksheet.cell(row=row, column=4)
                )
                worksheet.cell(row=row, column=5).value = ""
                row += 1

            end_row = row - 1

            worksheet.cell(row=start_row, column=1).value = student_number
            worksheet.cell(row=start_row, column=2).value = student

            if end_row > start_row:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=end_row,
                    end_column=1
                )

                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=2,
                    end_row=end_row,
                    end_column=2
                )

                if student_number % 2 == 1:
                    for r in range(start_row, end_row + 1):
                        for c in range(1, 6):   # Columnas A-E
                            worksheet.cell(row=r, column=c).fill = light_blue_fill

            student_number += 1

        worksheet.column_dimensions["A"].width = 5
        worksheet.column_dimensions["B"].width = 30
        worksheet.column_dimensions["C"].width = 140
        worksheet.column_dimensions["D"].width = 18
        worksheet.column_dimensions["E"].width = 40

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        course = self.cmb_course.get().replace(" ", "")
        filename = (
            f"{course}_"
            f"{self.cmb_module.get()}_"
            f"Term{self.cmb_term.get()}_"
            f"{self.cmb_course_language.get()}_"
            f"{timestamp}.xlsx"
        )

        output_folder = "output"
        os.makedirs(output_folder, exist_ok=True)
        filepath = os.path.join(output_folder, filename)
        filepath = os.path.abspath(filepath)

        workbook.save(filepath)

        students_count = len(students)
        questions_used = students_count * questions_per_student

        messagebox.showinfo(
            self.language.get("success_title"),
            f"{self.language.get('success_message')}\n\n"
            f"{self.language.get('file')}:\n"
            f"{filename}\n\n"
            f"{self.language.get('location')}:\n"
            f"{filepath}\n\n"
            f"{self.language.get('summary')}\n"
            f"────────────────────\n"
            f"{self.language.get('students')}: {students_count}\n"
            f"{self.language.get('questions_used')}: {questions_used}\n"
            f"{self.language.get('questions_per_student')}: {questions_per_student}"
        )

        print("Excel generated.")

    # =====================================================
    # Update Language
    # =====================================================
    def change_language(self, choice):

        language_code = self.config_data["lists"]["application_languages"][choice]

        self.language.load_language(language_code)

        self.title(
            f'{self.language.get("title")} v{self.config_data["application"]["version"]}'
        )

        self.lbl_title.configure(text=self.language.get("title"))
        self.lbl_application_language.configure(text=self.language.get("application_language"))
        self.lbl_course.configure(text=self.language.get("course"))
        self.lbl_module.configure(text=self.language.get("module"))
        self.lbl_term.configure(text=self.language.get("term"))
        self.lbl_year.configure(text=self.language.get("year"))
        self.lbl_course_language.configure(text=self.language.get("course_language"))
        self.lbl_students.configure(text=self.language.get("student_list"))
        self.lbl_questions.configure(text=self.language.get("question_bank"))
        self.btn_generate.configure(text=self.language.get("generate_excel"))
        self.btn_exit.configure(text=self.language.get("exit"))

        self.create_menu()
    
    def show_about(self):
        messagebox.showinfo(
            self.language.get("about"),
            "Question Assignment Tool\n\n"
            "Version 1.0\n\n"
            "Developed by\n"
            "Ing. Lucio M. Buitrón Pareja"
        )
    
    def show_instructions(self):

        messagebox.showinfo(
            self.language.get("instructions_title"),
            f"{self.language.get('instruction_1')}\n\n"
            f"{self.language.get('instruction_2')}\n\n"
            f"{self.language.get('instruction_3')}"
    )